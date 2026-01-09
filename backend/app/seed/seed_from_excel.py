"""
Excel Data Extraction and Database Seeding

Extracts reference data from the NVIDIA valuation Excel model:
- Country equity risk premiums
- Industry averages (US and Global)
- Synthetic rating tables
"""

import logging
from pathlib import Path
from decimal import Decimal
from typing import Optional

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Country, Industry, SyntheticRating

logger = logging.getLogger(__name__)

EXCEL_FILE = Path(__file__).parent.parent.parent.parent / "data" / "NvidiaJan2025.xlsx"


def safe_decimal(value, default: float = 0.0) -> Decimal:
    """Convert value to Decimal safely."""
    if value is None:
        return Decimal(str(default))
    try:
        if isinstance(value, str):
            value = value.strip().replace('%', '')
            if not value:
                return Decimal(str(default))
        return Decimal(str(float(value)))
    except (ValueError, TypeError):
        return Decimal(str(default))


def safe_int(value, default: int = 0) -> int:
    """Convert value to int safely."""
    if value is None:
        return default
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return default


def safe_string(value, default: str = "") -> str:
    """Convert value to string safely."""
    if value is None:
        return default
    return str(value).strip()


class ExcelDataExtractor:
    """Extracts data from the NVIDIA valuation Excel model."""

    def __init__(self, excel_path: Path = EXCEL_FILE):
        self.excel_path = excel_path
        self.workbook: Optional[openpyxl.Workbook] = None

    def open(self) -> None:
        """Open the Excel workbook."""
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")
        logger.info(f"Opening Excel file: {self.excel_path}")
        self.workbook = openpyxl.load_workbook(self.excel_path, data_only=True)

    def close(self) -> None:
        """Close the Excel workbook."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None

    def __enter__(self):
        self.open()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    def extract_countries(self) -> list[dict]:
        """
        Extract country equity risk premium data.

        Sheet: 'Country equity risk premiums'
        Columns: Country, Moody's rating, Adj Default Spread, Equity Risk Premium,
                 Country Risk Premium, Corporate Tax Rate
        """
        sheet_name = "Country equity risk premiums"
        if sheet_name not in self.workbook.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found")
            return []

        sheet = self.workbook[sheet_name]
        countries = []

        # Find header row and column indices
        header_row = None
        for row_idx in range(1, 10):
            cell_value = sheet.cell(row=row_idx, column=1).value
            if cell_value and "country" in str(cell_value).lower():
                header_row = row_idx
                break

        if header_row is None:
            logger.warning("Could not find header row in countries sheet")
            return []

        # Extract data starting from row after header
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            country_name = sheet.cell(row=row_idx, column=1).value
            if not country_name or str(country_name).strip() == "":
                continue

            country_data = {
                "name": safe_string(country_name),
                "moodys_rating": safe_string(sheet.cell(row=row_idx, column=2).value),
                "adj_default_spread": safe_decimal(sheet.cell(row=row_idx, column=3).value),
                "equity_risk_premium": safe_decimal(sheet.cell(row=row_idx, column=4).value),
                "country_risk_premium": safe_decimal(sheet.cell(row=row_idx, column=5).value),
                "corporate_tax_rate": safe_decimal(sheet.cell(row=row_idx, column=6).value),
            }
            countries.append(country_data)

        logger.info(f"Extracted {len(countries)} countries")
        return countries

    def extract_industries_us(self) -> list[dict]:
        """
        Extract US industry averages.

        Sheet: 'Industry Averages(US)'
        """
        return self._extract_industries("Industry Averages(US)", "US")

    def extract_industries_global(self) -> list[dict]:
        """
        Extract Global industry averages.

        Sheet: 'Industry Average Beta (Global)'
        """
        return self._extract_industries("Industry Average Beta (Global)", "Global")

    def _extract_industries(self, sheet_name: str, region: str) -> list[dict]:
        """Extract industry data from specified sheet."""
        if sheet_name not in self.workbook.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found")
            return []

        sheet = self.workbook[sheet_name]
        industries = []

        # Find header row
        header_row = None
        for row_idx in range(1, 10):
            cell_value = sheet.cell(row=row_idx, column=1).value
            if cell_value and "industry" in str(cell_value).lower():
                header_row = row_idx
                break

        if header_row is None:
            # Try looking for specific column headers
            for row_idx in range(1, 10):
                for col_idx in range(1, 5):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    if cell_value and ("industry" in str(cell_value).lower()
                                       or "sector" in str(cell_value).lower()):
                        header_row = row_idx
                        break
                if header_row:
                    break

        if header_row is None:
            logger.warning(f"Could not find header row in {sheet_name}")
            return []

        # Map column headers to indices
        col_map = {}
        for col_idx in range(1, sheet.max_column + 1):
            header = sheet.cell(row=header_row, column=col_idx).value
            if header:
                col_map[str(header).lower().strip()] = col_idx

        # Extract data
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            industry_name = sheet.cell(row=row_idx, column=1).value
            if not industry_name or str(industry_name).strip() == "":
                continue
            if "total" in str(industry_name).lower():
                continue

            industry_data = {
                "name": safe_string(industry_name),
                "region": region,
                "num_firms": safe_int(self._get_cell_by_header(sheet, row_idx, col_map,
                    ["number of firms", "# of firms", "firms", "num firms"])),
                "revenue_growth_5yr": safe_decimal(self._get_cell_by_header(sheet, row_idx, col_map,
                    ["revenue growth", "rev growth", "growth", "cagr"])),
                "operating_margin": safe_decimal(self._get_cell_by_header(sheet, row_idx, col_map,
                    ["operating margin", "op margin", "ebit margin"])),
                "after_tax_roc": safe_decimal(self._get_cell_by_header(sheet, row_idx, col_map,
                    ["after-tax roc", "roc", "roic", "return on capital"])),
                "effective_tax_rate": safe_decimal(self._get_cell_by_header(sheet, row_idx, col_map,
                    ["effective tax rate", "tax rate", "eff tax"])),
                "unlevered_beta": safe_decimal(self._get_cell_by_header(sheet, row_idx, col_map,
                    ["unlevered beta", "asset beta", "beta (unlevered)"])),
                "levered_beta": safe_decimal(self._get_cell_by_header(sheet, row_idx, col_map,
                    ["levered beta", "equity beta", "beta"])),
                "cost_of_equity": safe_decimal(self._get_cell_by_header(sheet, row_idx, col_map,
                    ["cost of equity", "coe", "ke"])),
                "std_dev_stock": safe_decimal(self._get_cell_by_header(sheet, row_idx, col_map,
                    ["std dev", "standard deviation", "volatility"])),
                "pretax_cost_of_debt": safe_decimal(self._get_cell_by_header(sheet, row_idx, col_map,
                    ["pre-tax cost of debt", "cost of debt", "pretax cod"])),
                "debt_to_capital": safe_decimal(self._get_cell_by_header(sheet, row_idx, col_map,
                    ["debt/capital", "d/(d+e)", "debt ratio", "debt to capital"])),
                "cost_of_capital": safe_decimal(self._get_cell_by_header(sheet, row_idx, col_map,
                    ["cost of capital", "wacc", "coc"])),
                "sales_to_capital": safe_decimal(self._get_cell_by_header(sheet, row_idx, col_map,
                    ["sales/capital", "sales to capital", "capital turnover"])),
                "ev_to_sales": safe_decimal(self._get_cell_by_header(sheet, row_idx, col_map,
                    ["ev/sales", "ev to sales"])),
                "ev_to_ebitda": safe_decimal(self._get_cell_by_header(sheet, row_idx, col_map,
                    ["ev/ebitda", "ev to ebitda"])),
                "ev_to_ebit": safe_decimal(self._get_cell_by_header(sheet, row_idx, col_map,
                    ["ev/ebit", "ev to ebit"])),
                "price_to_book": safe_decimal(self._get_cell_by_header(sheet, row_idx, col_map,
                    ["price/book", "p/b", "pbv", "price to book"])),
                "trailing_pe": safe_decimal(self._get_cell_by_header(sheet, row_idx, col_map,
                    ["trailing pe", "p/e", "pe ratio"])),
            }
            industries.append(industry_data)

        logger.info(f"Extracted {len(industries)} industries from {region}")
        return industries

    def _get_cell_by_header(self, sheet, row_idx: int, col_map: dict,
                            possible_headers: list[str]):
        """Get cell value by matching against possible header names."""
        for header in possible_headers:
            for col_name, col_idx in col_map.items():
                if header in col_name:
                    return sheet.cell(row=row_idx, column=col_idx).value
        return None

    def extract_synthetic_ratings(self) -> list[dict]:
        """
        Extract synthetic rating tables.

        Sheet: 'Synthetic rating'
        Contains ICR ranges mapped to ratings and spreads for different firm types.
        """
        sheet_name = "Synthetic rating"
        if sheet_name not in self.workbook.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found")
            return []

        sheet = self.workbook[sheet_name]
        ratings = []

        # The synthetic rating sheet typically has tables for:
        # 1. Large manufacturing firms (ICR > 8.5)
        # 2. Smaller/riskier firms
        # We need to find these tables and extract the data

        # Search for tables by looking for ICR ranges
        for row_idx in range(1, sheet.max_row + 1):
            for col_idx in range(1, min(20, sheet.max_column + 1)):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value and (">" in str(cell_value) or
                                   "to" in str(cell_value).lower() or
                                   "-" in str(cell_value)):
                    # This might be an ICR range
                    icr_text = str(cell_value).strip()

                    # Try to get rating and spread from adjacent cells
                    rating = sheet.cell(row=row_idx, column=col_idx + 1).value
                    spread = sheet.cell(row=row_idx, column=col_idx + 2).value

                    if rating and spread:
                        icr_min, icr_max = self._parse_icr_range(icr_text)
                        if icr_min is not None:
                            ratings.append({
                                "firm_type": "large_manufacturing",
                                "icr_min": icr_min,
                                "icr_max": icr_max,
                                "rating": safe_string(rating),
                                "spread": safe_decimal(spread),
                            })

        # If we didn't find data, use hardcoded values from original analysis
        if not ratings:
            ratings = self._get_default_synthetic_ratings()

        logger.info(f"Extracted {len(ratings)} synthetic rating entries")
        return ratings

    def _parse_icr_range(self, text: str) -> tuple[Optional[Decimal], Optional[Decimal]]:
        """Parse ICR range text like '> 12.5' or '8.5 to 12.5' or '6.5 - 8.5'."""
        text = text.strip()

        if text.startswith(">"):
            # Greater than (no upper bound)
            try:
                val = float(text.replace(">", "").strip())
                return Decimal(str(val)), Decimal("9999")
            except ValueError:
                return None, None

        if text.startswith("<"):
            # Less than (no lower bound)
            try:
                val = float(text.replace("<", "").strip())
                return Decimal("-9999"), Decimal(str(val))
            except ValueError:
                return None, None

        # Range with separator
        for sep in ["to", "-", "–"]:
            if sep in text.lower():
                parts = text.lower().split(sep)
                if len(parts) == 2:
                    try:
                        min_val = float(parts[0].strip())
                        max_val = float(parts[1].strip())
                        return Decimal(str(min_val)), Decimal(str(max_val))
                    except ValueError:
                        pass

        return None, None

    def _get_default_synthetic_ratings(self) -> list[dict]:
        """Return default synthetic rating data based on Damodaran's tables."""
        large_firm_ratings = [
            (12.5, 9999, "Aaa/AAA", 0.0063),
            (9.5, 12.5, "Aa2/AA", 0.0078),
            (7.5, 9.5, "A1/A+", 0.0098),
            (6.0, 7.5, "A2/A", 0.0108),
            (4.5, 6.0, "A3/A-", 0.0122),
            (4.0, 4.5, "Baa2/BBB", 0.0156),
            (3.5, 4.0, "Ba1/BB+", 0.0200),
            (3.0, 3.5, "Ba2/BB", 0.0240),
            (2.5, 3.0, "B1/B+", 0.0324),
            (2.0, 2.5, "B2/B", 0.0404),
            (1.5, 2.0, "B3/B-", 0.0528),
            (1.25, 1.5, "Caa/CCC", 0.0827),
            (0.8, 1.25, "Ca2/CC", 0.0877),
            (0.5, 0.8, "C2/C", 0.1152),
            (-9999, 0.5, "D2/D", 0.1502),
        ]

        small_firm_ratings = [
            (8.5, 9999, "Aaa/AAA", 0.0063),
            (6.5, 8.5, "Aa2/AA", 0.0078),
            (5.5, 6.5, "A1/A+", 0.0098),
            (4.25, 5.5, "A2/A", 0.0108),
            (3.0, 4.25, "A3/A-", 0.0122),
            (2.5, 3.0, "Baa2/BBB", 0.0156),
            (2.25, 2.5, "Ba1/BB+", 0.0200),
            (2.0, 2.25, "Ba2/BB", 0.0240),
            (1.75, 2.0, "B1/B+", 0.0324),
            (1.5, 1.75, "B2/B", 0.0404),
            (1.25, 1.5, "B3/B-", 0.0528),
            (0.8, 1.25, "Caa/CCC", 0.0827),
            (0.65, 0.8, "Ca2/CC", 0.0877),
            (0.2, 0.65, "C2/C", 0.1152),
            (-9999, 0.2, "D2/D", 0.1502),
        ]

        ratings = []
        for icr_min, icr_max, rating, spread in large_firm_ratings:
            ratings.append({
                "firm_type": "large_manufacturing",
                "icr_min": Decimal(str(icr_min)),
                "icr_max": Decimal(str(icr_max)),
                "rating": rating,
                "spread": Decimal(str(spread)),
            })

        for icr_min, icr_max, rating, spread in small_firm_ratings:
            ratings.append({
                "firm_type": "small_risky",
                "icr_min": Decimal(str(icr_min)),
                "icr_max": Decimal(str(icr_max)),
                "rating": rating,
                "spread": Decimal(str(spread)),
            })

        return ratings


async def seed_countries(db: AsyncSession, countries: list[dict]) -> int:
    """Seed country data to database."""
    count = 0
    for country_data in countries:
        # Check if country exists
        result = await db.execute(
            select(Country).where(Country.name == country_data["name"])
        )
        existing = result.scalar_one_or_none()

        if existing:
            # Update existing
            for key, value in country_data.items():
                setattr(existing, key, value)
        else:
            # Create new
            country = Country(**country_data)
            db.add(country)
            count += 1

    await db.commit()
    logger.info(f"Seeded {count} new countries")
    return count


async def seed_industries(db: AsyncSession, industries: list[dict]) -> int:
    """Seed industry data to database."""
    count = 0
    for industry_data in industries:
        # Check if industry exists
        result = await db.execute(
            select(Industry).where(
                Industry.name == industry_data["name"],
                Industry.region == industry_data["region"]
            )
        )
        existing = result.scalar_one_or_none()

        if existing:
            # Update existing
            for key, value in industry_data.items():
                setattr(existing, key, value)
        else:
            # Create new
            industry = Industry(**industry_data)
            db.add(industry)
            count += 1

    await db.commit()
    logger.info(f"Seeded {count} new industries")
    return count


async def seed_synthetic_ratings(db: AsyncSession, ratings: list[dict]) -> int:
    """Seed synthetic rating data to database."""
    count = 0
    for rating_data in ratings:
        # Check if rating exists
        result = await db.execute(
            select(SyntheticRating).where(
                SyntheticRating.firm_type == rating_data["firm_type"],
                SyntheticRating.rating == rating_data["rating"]
            )
        )
        existing = result.scalar_one_or_none()

        if existing:
            # Update existing
            for key, value in rating_data.items():
                setattr(existing, key, value)
        else:
            # Create new
            rating = SyntheticRating(**rating_data)
            db.add(rating)
            count += 1

    await db.commit()
    logger.info(f"Seeded {count} new synthetic ratings")
    return count


async def seed_all_from_excel(db: AsyncSession, excel_path: Path = EXCEL_FILE) -> dict:
    """
    Extract all reference data from Excel and seed to database.

    Returns summary of seeded data counts.
    """
    logger.info("Starting Excel data extraction and seeding")

    with ExcelDataExtractor(excel_path) as extractor:
        # Extract data
        countries = extractor.extract_countries()
        industries_us = extractor.extract_industries_us()
        industries_global = extractor.extract_industries_global()
        ratings = extractor.extract_synthetic_ratings()

    # Seed to database
    countries_count = await seed_countries(db, countries)
    industries_us_count = await seed_industries(db, industries_us)
    industries_global_count = await seed_industries(db, industries_global)
    ratings_count = await seed_synthetic_ratings(db, ratings)

    summary = {
        "countries": countries_count,
        "industries_us": industries_us_count,
        "industries_global": industries_global_count,
        "synthetic_ratings": ratings_count,
    }

    logger.info(f"Seeding complete: {summary}")
    return summary
